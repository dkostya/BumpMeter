# coding=utf-8
import serial
from datetime import datetime
import winsound
import openpyxl
import os
from openpyxl.chart import LineChart, Reference

# Основные переменные

accel_list = []

secondsInMem = 5

result_range = 100

file_name = 'results.xlsx'

file_raw_name = 'raw_data.xlsx'

shot_name = input('Введи название:')

# открываем порт
ser = serial.Serial('COM4', 115200)

print(ser)

for j in range(10):
    print(ser.readline())

# процедура чтения и декодирования строки из последовательного порта
def mySerialDecode():
    accel = (((ser.readline()
               .decode('utf-8', 'ignore')))
               .strip())\
               .split(',')
    accel.pop() #удаляем последний элемент, это запятая!
    accel_values = [int(x) for x in accel]
    return accel_values


# Запись данных в Excel
def dataToExcel(data, file_name):

    if os.path.exists(file_name):  # Если файл существует
        wb = openpyxl.load_workbook(file_name)
        ws = wb['Res']
        next_col = ws.max_column + 1

        for i in range(len(data)):

            ws.cell(row=i+1, column=next_col).value = data[i]

        wb.save(file_name)

    else: # Файла нет, создаем

        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'Res'

        for i in range(len(data)):

            ws1.cell(row=i+1, column=1).value = data[i]

        wb.save(file_name)


start_time = datetime.now()  # Timer start

# Сбор показаний
for i in range(secondsInMem * 100):

    if i % 10 == 0:
        print(i*10)

    if  i == (secondsInMem * 100 // 3):
        winsound.Beep(2500, 400)



    accel_list.extend((mySerialDecode()))

print('Время сбора данных: ', datetime.now() - start_time)

# Удаляю первый элемент списка, т.к. в большинстве случаев он ошибочный
accel_list.pop(0)

# Поиск и проверка максимума.
max_force = accel_list.index(max(accel_list))


# Проверка максимума. Бывают непонятные выбросы
if (accel_list[max_force] // accel_list[max_force-2]) > 5:
    del accel_list[max_force-3:max_force+3]
    max_force = accel_list.index(max(accel_list))


# Создание результирующего списка значений
bump_plot = [shot_name] # Заголовок столбца с данными

# Запись отрезка с найденным выстрелом
if max_force - result_range < 0:
    bump_plot.extend(accel_list[:max_force])
else:
    bump_plot.extend(accel_list[max_force - result_range:max_force])
bump_plot.extend(accel_list[max_force:max_force + result_range + 50])



# Вывод показаний в консоль
print('Полный список значений: ', accel_list)
print('Количество значений', len(accel_list))
print('Максимальное значение: ', accel_list[max_force])
print('Выборка: ', bump_plot)
print('Количество значений выборки', len(bump_plot))

# Записываю данные в excel
dataToExcel(bump_plot, file_name)
dataToExcel(accel_list, file_raw_name)

# А здесь попробуем нарисовать красивые графики

wb = openpyxl.load_workbook(file_name)
raw_wb = openpyxl.load_workbook(file_raw_name) # файл с сырыми данными

ws = wb['Res']
raw_ws = raw_wb['Res']

chart = LineChart()
data = Reference(ws, min_col=1, min_row=1, max_col=ws.max_column, max_row=ws.max_row)
chart.add_data(data, titles_from_data=True)

raw_chart = LineChart()
raw_data = Reference(raw_ws, min_col=1, min_row=1, max_col=raw_ws.max_column, max_row=raw_ws.max_row)
raw_chart.add_data(raw_data, titles_from_data=False)

# Сглаживание линий графика
for line in chart.series:

    line.smooth = True

for raw_line in raw_chart.series:

    raw_line.smooth = True


# Создаем график на странице
ws.add_chart(chart, "A10")

wb.save(file_name)

raw_ws.add_chart(raw_chart, "A10")

raw_wb.save(file_raw_name)

if True:
    print('done')
